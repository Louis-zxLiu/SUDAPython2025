import openpyxl as px
import random

def safe_int(value):
    try:
        return int(value)
    except:
        return 0

class TxtFile:
    @classmethod
    def getGameInfo(cls):
        try:
            with open("小测验游戏说明.txt", "r", encoding="gbk") as f:
                return f.read()
        except FileNotFoundError:
            return "错误：未找到小测验游戏说明.txt文件！"

    @classmethod
    def getMaxScore(cls):
        try:
            with open("Record.txt", "r", encoding="utf-8") as f:
                lines = f.readlines()
                if len(lines) >= 3:
                    return int(lines[2].split(":")[1].strip())
        except:
            pass
        return -1

    @classmethod
    def setNewRecord(cls, sno, sname, score):
        with open("Record.txt", "w", encoding="utf-8") as f:
            f.write(f"学号: {sno}\n")
            f.write(f"姓名: {sname}\n")
            f.write(f"得分: {score}\n")

class Stu:
    def __init__(self):
        try:
            self.wb = px.load_workbook("名单.xlsx", data_only=True)
            self.sheet = self.wb.active
        except Exception as e:
            raise FileNotFoundError(f"加载名单.xlsx失败: {str(e)}")

    def getstu(self):
        students = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):  # 从第2行开始
            if len(row) < 2 or row[0] is None or row[1] is None:
                continue
                
            sno = str(row[0]).strip()
            name = str(row[1]).strip()
            
            # 验证学号是否为10位数字
            if not (sno.isdigit() and len(sno) == 10):
                continue
                
            if sno and name:
                students.append((sno, name))
        
        # 确保至少有7个学生
        if len(students) < 7:
            print(f"警告: 只找到 {len(students)} 个有效学生记录")
        
        return students

class QusAndAns:
    def __init__(self):
        try:
            self.wb = px.load_workbook("Quiz.xlsx")
            self.sheet_info = self.wb["测验信息"]
            self.sheet_select = self.wb["选择题"]
            self.sheet_fill = self.wb["填空题"]
            self.sheet_tf = self.wb["判断题"]
        except Exception as e:
            raise e

        self.examName = self.sheet_info["B1"].value or "未命名测验"
        self.examTime = safe_int(self.sheet_info["B2"].value)

        self.selectScore = safe_int(self.sheet_select["F1"].value)
        self.fillScore = safe_int(self.sheet_fill["F1"].value)
        self.tfScore = safe_int(self.sheet_tf["F1"].value)

        self.questions = []

    def getExamNameAndTime(self):
        return self.examName, self.examTime

    def getQuestions(self):
        # 选择题处理
        for row in self.sheet_select.iter_rows(min_row=3, max_col=9, values_only=True):
            if not row[1]:
                continue
            question = {
                "type": "select",
                "text": row[1],
                "options": [row[2], row[3], row[4], row[5]],
                "answer": str(row[6]).strip().upper()[0],
                "explanation": row[7] or "",
                "score": self.selectScore
            }
            # 打乱选项
            indices = list(range(4))
            random.shuffle(indices)
            shuffled_options = [question["options"][i] for i in indices]
            original_index = ord(question["answer"]) - ord("A")
            new_index = indices.index(original_index)
            question["options"] = shuffled_options
            question["answer"] = chr(ord("A") + new_index)
            self.questions.append(question)

        # 填空题处理
        for row in self.sheet_fill.iter_rows(min_row=3, max_col=6, values_only=True):
            if not row[1]:
                continue
            self.questions.append({
                "type": "fill",
                "text": row[1],
                "answer": str(row[2]).strip(),
                "explanation": row[3] or "",
                "score": self.fillScore
            })

        # 判断题处理
        for row in self.sheet_tf.iter_rows(min_row=3, max_col=6, values_only=True):
            if not row[1]:
                continue
            self.questions.append({
                "type": "tf",
                "text": row[1],
                "answer": str(row[2]).strip(),
                "explanation": row[3] or "",
                "score": self.tfScore
            })

        random.shuffle(self.questions)
        return self.questions