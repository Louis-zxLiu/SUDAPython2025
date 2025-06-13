mWin=0
tno=0
tname=0 
note=0
# tqus=0
blogin=0 
L_info=0 
bnext=0 
v=0

Lresult=0
# tans=0 
ba=0
bb=0
bc=0
bd=0

# 添加收藏按钮的全局变量
bfavorite=0
bfavorite_blank=0
bfavorite_pd=0

# 添加错题记录相关的全局变量
favorite_questions = []  # 存储收藏的题目

# 添加全局变量
consecutive_correct = 0  # 连续答对次数
stars_frame = None  # 星星容器
star_labels = []  # 星星标签列表

from Exam import *
qusAndAns=QusAndAns()
qus=qusAndAns.getQusOfSelect()
ans=qusAndAns.getAnsOfSelect()
blanks=qusAndAns.getQusOfBlank()
blank_ans=qusAndAns.getAnsOfBlank()
analyze=qusAndAns.getAnalyzeOfSelect()

pd=qusAndAns.getQusOpd()
pd_ans=qusAndAns.getAnsOpd()
randpd=qusAndAns.getRandQusOfpd()


eName,eTime=qusAndAns.getEnameAndEtime()
randselect=qusAndAns.getRandQusOfSelect()
totalSelect,iselectScore=qusAndAns.getTotalAndiScore()
totalScore=0 
logstate=0 
indexofselect=1
indexofpd=1
# 登录调用函数，验证学号和姓名后初始化游戏
def loginCall():
    global logstate
    sno=tno.get()
    sname=tname.get()
    stu=Stu()
    stu_info=stu.getStu()
    if(sno,sname)in stu_info:
        logstate=1
        mb.showinfo("QuizGame","登录成功，游戏启动！")
        note.tab(1,state="normal")
        tqus.config(state="normal")
        tqus.focus_set()
        tno.config(state="disable")
        tname.config(state="disable")
        blogin['state']="disable"
        tabl = note. tabs()[1]
        note.select(tabl)
        Tl=myTimer(mWin,L_info,eTime*60)
        Tl.start()
    else:
        response=mb.showinfo("QuizGame","学号或姓名错误，您要注册吗？")
        if response:
            register()
def register():
    # Create a new window for registration
    reg_window = tk.Toplevel(mWin)
    reg_window.title("注册")
    reg_window.geometry("300x200")
     # Create labels and entry widgets for student number and name
    lno_reg = ttk.Label(reg_window, text="学号：")
    lname_reg = ttk.Label(reg_window, text="姓名：")
    tno_reg = ttk.Entry(reg_window)
    tname_reg = ttk.Entry(reg_window)
    bregister = ttk.Button(reg_window, text="注册", command=lambda: save_registration(tno_reg.get(), tname_reg.get()))
    # Layout the labels and entry widgets
    lno_reg.grid(column=0, row=0, padx=10, pady=10)
    lname_reg.grid(column=0, row=1, padx=10, pady=10)
    tno_reg.grid(column=1, row=0, padx=10, pady=10)
    tname_reg.grid(column=1, row=1, padx=10, pady=10)
    bregister.grid(column=1, row=2, padx=10, pady=10)

def save_registration(sno, sname):
    # Open the Excel file and add the new student details
    wb = px.load_workbook('名单.xlsx')
    sheet = wb.active
    max_row = sheet.max_row
    sheet.cell(row=max_row + 1, column=1).value = sno
    sheet.cell(row=max_row + 1, column=2).value = sname
    wb.save('名单.xlsx')
    mb.showinfo("注册成功", "学号和姓名已添加到名单中。")

    
def click(event):
    global logstate
    if logstate==0:
        mb.showinfo("QuizGame","请先登录后开始测验游戏！")

def load(event):
    global indexofselect
    tqus.insert("0.0", str(indexofselect)+qus[randselect[0]])
    tqus.config(state="disable")

def play_sound_thread(sound_type):
    """在新线程中播放音效
    Args:
        sound_type: 'correct' 或 'wrong'
    """
    def play():
        sound_file = os.path.join('sounds', f'{sound_type}.wav')
        if os.path.exists(sound_file):
            winsound.PlaySound(sound_file, winsound.SND_FILENAME)
        else:
            # 使用Windows系统默认音效
            if sound_type == 'correct':
                winsound.PlaySound('SystemExclamation', winsound.SND_ALIAS)
            else:
                winsound.PlaySound('SystemHand', winsound.SND_ALIAS)
    
    # 创建并启动新线程播放音效
    sound_thread = threading.Thread(target=play)
    sound_thread.daemon = True  # 设置为守护线程，这样主程序退出时线程会自动结束
    sound_thread.start()

def radioCall():
    global indexofselect, totalScore, consecutive_correct
    if v.get() == "E":  # 检查是否选择了答案
        mb.showinfo("提示", "您还没有选择答案！")
        return
    if(v.get()!=ans[randselect[indexofselect-1]]):
        Lresult.configure(text='错误')
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "red")])
        tans.config(state='normal')
        tans.delete("0.0", "end")  # 清空答案文本框
        s="答案："+ans[randselect[indexofselect-1]]+"\n"+"解析："+analyze[randselect[indexofselect-1]]
        tans.insert("0.0",s)
        tans.config(state="disable")
        play_sound_thread('wrong')  # 在新线程中播放错误音效
        reset_stars()  # 答错时重置星星
        # 更新收藏按钮文本
        for q in favorite_questions:
            if q["type"] == "选择题" and q["question"] == qus[randselect[indexofselect-1]]:
                bfavorite.config(text="移除收藏")
                break
        else:
            bfavorite.config(text="加入收藏")
    else:
        Lresult.configure(text="正确")
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "green")])
        totalScore+=iselectScore
        consecutive_correct += 1  # 增加连续答对次数
        update_stars()  # 更新星星显示
        play_sound_thread('correct')  # 在新线程中播放正确音效
    ba.config(state="disable")
    bb.config(state="disable")
    bc.config(state="disable")
    bd.config(state="disable")

def nextCall():
    global indexofselect,T1
    global v
    if v.get() == "E":  # 检查是否选择了答案
        mb.showinfo("提示", "您还没有选择答案！")
        return
    v.set("E")
    indexofselect+=1
    Lresult.config(text="")
    if(indexofselect>len(qus)):
        indexofselect=len(qus) 
        bnext.config(state="disable")
        mb.showinfo("QuizGame","选择题结束，进入填空题！")
        note.tab(2,state="normal")
        tab2 = note.tabs()[2]
        note.select(tab2)
        load_blank_question()
    ba.config(state="normal")
    bb.config(state="normal")
    bc.config(state="normal")
    bd.config(state="normal")
    tqus.config(state="normal")
    tans.config(state="normal")
    tqus.delete("0.0", "40.200")
    tans.delete("0.0", "40.200")
    tqus.insert('0.0',str(indexofselect)+qus[randselect[indexofselect-1]])
    tqus.config(state="disable")
    tans.config(state="disable")
    # 更新收藏按钮文本
    for q in favorite_questions:
        if q["type"] == "选择题" and q["question"] == qus[randselect[indexofselect-1]]:
            bfavorite.config(text="移除收藏")
            break
    else:
        bfavorite.config(text="加入收藏")

def nextCallpd():
    global indexofpd, totalScore, vv, tqus1, pd, pd_ans, Lresult1, tans1, bnext1
    if vv.get() == "E":  # 检查是否选择了答案
        mb.showinfo("提示", "您还没有选择答案！")
        return
    vv.set("E")
    indexofpd += 1
    Lresult1.config(text="")
    if indexofpd > len(pd):
        GameOver()
    else:
        ba1.config(state="normal")
        bb1.config(state="normal")
        tqus1.config(state="normal")
        tans1.config(state="normal")
        tqus1.delete("0.0", "end")
        tans1.delete("0.0", "end")
        tqus1.insert("0.0", str(indexofpd) + ". " + pd[randpd[indexofpd-1]])
        tqus1.config(state="disable")
        tans1.config(state="disable")

def prev_question():
    global indexofselect
    if indexofselect > 1:
        indexofselect -= 1
    question = qus[randselect[indexofselect - 1]]
    tqus.config(state="normal")
    tqus.delete("0.0", "end")
    tqus.insert("0.0", str(indexofselect) + ". " + question)
    tqus.config(state="disable")
    Lresult.config(text="")
    ba.config(state="normal")
    bb.config(state="normal")
    bc.config(state="normal")
    bd.config(state="normal")

def prev_pd():
    global indexofpd
    if indexofpd > 1:
        indexofpd -= 1
    question = pd[randpd[indexofpd - 1]]
    tqus1.config(state="normal")
    tqus1.delete("0.0", "end")
    tqus1.insert("0.0", str(indexofpd) + ". " + question)
    tqus1.config(state="disable")
    Lresult1.config(text="")
    ba1.config(state="normal")
    bb1.config(state="normal")
   

        
        
# Define function to load the fill-in-the-blank question
def load_blank_question():
    # Load fill-in-the-blank questions
    global blank_index, blanks, blank_ans
    blank_index = 0
    tqus_blank.config(state="normal")
    tqus_blank.insert("0.0", str(blank_index + 1) + blanks[blank_index])
    tqus_blank.config(state="disable")
def load_pd():
    global indexofpd, pd, pd_ans
    indexofpd = 1  # 初始化为1而不是0，因为indexofpd是从1开始的
    tqus1.config(state="normal")
    tqus1.insert("0.0", str(indexofpd) + ". " + pd[randpd[indexofpd-1]])
    tqus1.config(state="disable")
    


def submit_blank_answer():
    global blank_index, totalScore, blank_ans, blanks, consecutive_correct
    user_answer = tanswer.get()
    if not user_answer.strip():  # 检查答案是否为空
        mb.showinfo("提示", "您还没有填写答案！")
        return
    if user_answer == blank_ans[blank_index]:
        totalScore += 10  # Assuming each fill-in-the-blank question is worth 10 points
        Lresult_blank.configure(text="正确")
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "green")])
        consecutive_correct += 1  # 增加连续答对次数
        update_stars()  # 更新星星显示
        play_sound_thread('correct')  # 在新线程中播放正确音效
    else:
        Lresult_blank.configure(text="错误")
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "red")])
        tans_blank.config(state="normal")
        tans_blank.delete("0.0", "end")  # 清空答案文本框
        s = "答案：" + blank_ans[blank_index] + "\n"
        tans_blank.insert("0.0", s)
        tans_blank.config(state="disable")
        reset_stars()  # 答错时重置星星
        play_sound_thread('wrong')  # 在新线程中播放错误音效
        # 更新收藏按钮文本
        for q in favorite_questions:
            if q["type"] == "填空题" and q["question"] == blanks[blank_index]:
                bfavorite_blank.config(text="移除收藏")
                break
        else:
            bfavorite_blank.config(text="加入收藏")
    blank_index += 1
    if blank_index >= len(blanks):
        mb.showinfo("QuizGame", "填空题结束，进入判断题！")
        note.tab(3,state="normal")
        tab3 = note.tabs()[3]
        note.select(tab3)
        load_pd()
    else:
        tanswer.delete(0, tk.END)
        tqus_blank.config(state="normal")
        tqus_blank.delete("0.0", "40.200")
        tqus_blank.insert("0.0", str(blank_index + 1) + blanks[blank_index])
        tqus_blank.config(state="disable")

def radioCa():
    global indexofpd, totalScore, consecutive_correct
    if vv.get() == "E":  # 检查是否选择了答案
        mb.showinfo("提示", "您还没有选择答案！")
        return
    if indexofpd > len(randpd):  # 添加边界检查
        GameOver()
        return
    if(vv.get()!=pd_ans[randpd[indexofpd-1]]):
        Lresult1.configure(text='错误')
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "red")])
        tans1.config(state='normal')
        tans1.delete("0.0", "end")  # 清空答案文本框
        s="答案："+pd_ans[randpd[indexofpd-1]]
        tans1.insert("0.0",s)
        tans1.config(state="disable")
        reset_stars()  # 答错时重置星星
        play_sound_thread('wrong')  # 在新线程中播放错误音效
        # 更新收藏按钮文本
        for q in favorite_questions:
            if q["type"] == "判断题" and q["question"] == pd[randpd[indexofpd-1]]:
                bfavorite_pd.config(text="移除收藏")
                break
        else:
            bfavorite_pd.config(text="加入收藏")
    else:
        Lresult1.configure(text="正确")
        style = ttk.Style()
        style.map("Result.TLabel", foreground=[("", "green")])
        totalScore+=iselectScore
        consecutive_correct += 1  # 增加连续答对次数
        update_stars()  # 更新星星显示
        play_sound_thread('correct')  # 在新线程中播放正确音效
    ba1.config(state="disable")
    bb1.config(state="disable")
    
    
    
    
def GameOver():
    result=TxtFile.getMaxScore()
    if result==-1 or totalScore>result:
        mb.showinfo(title="QuizGame", message="您的分数为：%d\n恭喜您创造了新的得分记录！\n我们将记录你的成绩！谢谢！"%totalScore)
        sno=tno.get()
        sname=tname.get()
        TxtFile.setNewRecord(sno,sname,totalScore)
    else:
        mb.showinfo(title="QuizGame", message="您的分数为：%d\n对不起，你没有创造新的得分记录！\n下次再努力哦！谢谢！"%totalScore)
    # 不再关闭程序
    # mWin.destroy()
    # sys.exit()

# 导入Tkinter模块，用于创建图形用户界面
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as mb
import tkinter.scrolledtext as sc
import openpyxl as px
import sys
import datetime
import time
import os
import winsound
import threading
from Exam import *

# 定义一个计时器类，用于在GUI中显示倒计时
class myTimer:
    # 初始化计时器对象，设置窗口、标签、持续时间、超时时调用的函数
    def __init__(self,window,label,seconds):
        self.hours=0
        self.minutes=0
        self.seconds=0
        self.is_running=False # 计时器是否正在运行的标志
        self.start_time=None
        self.window=window
        self.label=label
        self.duration=seconds
    # 将秒数格式化为分钟和秒的格式    
    def format_time(self,seconds):
        self.hours=seconds//3600
        self.minutes=(seconds%3600)//60
        self.seconds=seconds%60
        return "%02d:%02d:%02d"%(self.hours, self.minutes, self.seconds)
    # 更新时间显示的函数
    def update_time(self):
        if self.is_running:
            remaining_time = self.duration - (datetime.datetime.now()-self.start_time).total_seconds()
        if remaining_time<=0:
            self.is_running = False
            self.label.config(text="剩余时间："+self.format_time(0))
            GameOver()
        else:
            t=self.format_time(int(remaining_time))
            self.label.config(text="学号：%s   姓名：%s   剩余时间：%s"%(tno.get(),tname.get(),t))
            self.window.after(1000, self.update_time)
      # 开始计时器       
    def start(self):
        self.start_time = datetime.datetime.now()
        self.is_running = True
        self.update_time()
# 停止计时器
    def stop(self):
        self.is_running = False
        self.format_time(0)

def add_to_favorites(question_type, question, answer, analysis=""):
    """添加题目到收藏夹"""
    question_data = {
        "type": question_type,
        "question": question,
        "answer": answer,
        "analysis": analysis
    }
    if question_data not in favorite_questions:
        favorite_questions.append(question_data)
        update_favorites_display()
        return True
    return False

def remove_from_favorites(question_type, question):
    """从收藏夹移除题目"""
    for q in favorite_questions:
        if q["type"] == question_type and q["question"] == question:
            favorite_questions.remove(q)
            update_favorites_display()
            return True
    return False

def update_favorites_display():
    """更新收藏夹显示"""
    favorite_text.config(state="normal")
    favorite_text.delete("0.0", "end")
    for i, q in enumerate(favorite_questions, 1):
        favorite_text.insert("end", f"第{i}题 ({q['type']})\n")
        favorite_text.insert("end", f"题目：{q['question']}\n")
        favorite_text.insert("end", f"答案：{q['answer']}\n")
        if q['analysis']:
            favorite_text.insert("end", f"解析：{q['analysis']}\n")
        favorite_text.insert("end", "-" * 50 + "\n")
    favorite_text.config(state="disabled")

def toggle_favorite(question_type, question, answer, analysis=""):
    """切换收藏状态"""
    for q in favorite_questions:
        if q["type"] == question_type and q["question"] == question:
            remove_from_favorites(question_type, question)
            return False
    add_to_favorites(question_type, question, answer, analysis)
    return True

def update_stars():
    """更新星星显示"""
    global consecutive_correct, stars_frame, star_labels
    
    # 清除所有星星
    for label in star_labels:
        label.destroy()
    star_labels.clear()
    
    # 计算应该显示的星星数量
    num_stars = consecutive_correct // 5
    
    # 创建新的星星
    for i in range(num_stars):
        star_label = ttk.Label(stars_frame, text="★", style="Star.TLabel")
        star_label.pack(side=tk.TOP, pady=5)
        star_labels.append(star_label)

def reset_stars():
    """重置星星显示"""
    global consecutive_correct, star_labels
    consecutive_correct = 0
    for label in star_labels:
        label.destroy()
    star_labels.clear()

def main():
    # 使用 global 关键字声明，使得以下变量在函数内部可修改，它们在函数外也被定义
    global mWin,tno,tname,note,tqus,blogin,L_info,bnext,bprev,v,vv,Lresult,tans,ba,bb,bc,bd,blank_index, tqus_blank, tanswer, bsubmit_blank, Lresult_blank, tans_blank, blanks, blank_ans,bnext1,bprev1,Lans1,tans1,tqus1,Lresult1,Lqus1,Lans1,ba1,bb1,pd_index,tqus_pd, pd, pd_ans, favorite_text, bfavorite, bfavorite_blank, bfavorite_pd, stars_frame, star_labels
    # 创建一个 Tkinter 主窗口实例，并赋值给全局变量 mWin
    mWin=tk.Tk()
    # 获取屏幕的宽度和高度，用于计算窗口的起始位置
    screen_width = mWin.winfo_screenwidth()
    screen_height = mWin.winfo_screenheight()
    # 设置窗口的尺寸
    window_width = 1000
    window_height = 800
    # 计算窗口的 x 和 y 起始坐标，使得窗口在屏幕中居中
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    # 设置窗口的尺寸和位置
    mWin.geometry(f"{window_width}x{window_height}+{x}+{y}")
    # 禁止用户调整窗口大小
    mWin.resizable(width=False, height=False)
    # 设置窗口标题，显示游戏名称、测验名称和持续时间
    mWin.title("欢迎使用QuizGame"+""*50+"测验名称："+eName+""*10+"测验时长（分）:"+str(eTime))

    # 设置全局样式
    style = ttk.Style()
    style.configure("TLabel", font=("宋体", 12))
    style.configure("TButton", font=("宋体", 12))
    style.configure("TEntry", font=("宋体", 12))
    # 添加结果标签的样式
    style.configure("Result.TLabel", font=("宋体", 30, "bold"))
    style.configure("Result.TLabel.Correct", foreground="green")
    style.configure("Result.TLabel.Wrong", foreground="red")
    # 添加星星标签的样式
    style.configure("Star.TLabel", font=("宋体", 24), foreground="gold")

    # 创建主框架
    main_frame = ttk.Frame(mWin)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # 创建左侧内容框架
    content_frame = ttk.Frame(main_frame)
    content_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # 创建右侧星星框架
    stars_frame = ttk.Frame(main_frame, width=100)
    stars_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)
    stars_frame.pack_propagate(False)  # 固定宽度

    # 创建一个 Label 控件用于显示剩余时间，并放置在窗口底部
    L_info=ttk.Label(content_frame, text="剩余时间：")
    L_info.pack(side=tk.BOTTOM)

    # 创建一个 Notebook 控件，用于在不同页面间进行切换，并填充整个窗口
    note=ttk.Notebook(content_frame)
    note.pack(fill="both",expand=True)
    # 创建登录页面的 Frame 容器，并添加到 Notebook 控件中
    frml=ttk.Frame(note)
    frml.pack(fill="both",expand=True)
    note.add(frml,text="登录")
    # 在登录页面中创建一个 ScrolledText 控件，用于显示游戏说明
    textl=sc.ScrolledText (frml,height=30,width=100,wrap=tk.WORD)
    # 在 ScrolledText 控件中插入游戏说明的标题和内容
    textl.insert("0.0","\n\n"+"QuizGame游戏说明".center(100))
    textl.insert("25.0 linestart", "\n\n\n\n"+TxtFile.getGameInfo())
    # 创建学号和姓名的 Label 和 Entry 控件，以及登录按钮，并设置它们在页面中的布局
    lno=ttk.Label (frml, text="学号：")
    lname=ttk.Label (frml, text="姓名：")
    tno=ttk.Entry(frml)
    tname=ttk.Entry(frml)
    blogin=ttk.Button(frml,text="登录",command=loginCall)

    # 使用 grid 布局管理器对登录页面的控件进行布局
    textl.grid(column=0, row=0, columnspan=8, rowspan=30, padx=50,pady=20,sticky=tk.E)
    lno.grid(column=3,row=30,padx=2,pady=5,sticky=tk.E)
    tno.grid(column=4,row=30,padx=2,pady=5,sticky=tk.W) 
    lname.grid(column=3,row=31,padx=2, pady=5, sticky=tk.E)
    tname.grid(column=4,row=31,padx=2,pady=5, sticky=tk. W)
    blogin.grid(column=4,row=32,padx=5,pady=5,sticky=tk.W)
    # 创建选择题页面的 Frame 容器，并添加到 Notebook 控件中
    frm2=ttk.Frame(note)
    frm2.pack(fill="both",expand=True)
    note.add(frm2, text="选择题",state="disabled")
    # 在选择题页面中创建 Label 和 ScrolledText 控件用于显示题目，以及 Radiobutton 控件用于选择答案
    Lqus=ttk.Label(frm2,text="题目：")
    tqus=sc.ScrolledText(frm2)
    tqus.config(state="disabled")
    v=tk.StringVar()# 创建一个字符串变量，用于跟踪 Radiobutton 的选择
    ba=ttk.Radiobutton(frm2, text="A", variable=v, value="A",command=radioCall)
    bb=ttk.Radiobutton(frm2, text="B", variable=v, value="B",command=radioCall)
    bc=ttk.Radiobutton(frm2, text="C", variable=v, value="C",command=radioCall)
    bd=ttk.Radiobutton(frm2, text="D", variable=v, value="D",command=radioCall)
    # 创建下一题和上一题的按钮，以及解析 Label 和 ScrolledText 控件
    bnext=ttk.Button(frm2,text="下一题",command=nextCall)
    bprev = ttk.Button(frm2, text="上一题", command=prev_question)
    Lans=ttk.Label(frm2,text="解析：")
    tans=sc.ScrolledText(frm2) 
    tans.config(state="disabled")
    # 为题目文本框添加当获得焦点时自动加载题目的事件绑定
    tqus.bind("<FocusIn>",load)
    # 创建用于显示用户答题结果的 Label 控件
    Lresult=tk.Label(frm2)
    # 使用 grid 布局管理器对选择题页面的控件进行布局
    Lqus.grid(column=0, row=1, padx=20,pady=15, sticky=tk.W)
    Lans.grid(column=9, row=1, padx=5, pady=15,sticky=tk.W)
    tqus.config(height=30,width=80)
    tqus.grid(column=0, row=2, columnspan=9, rowspan=30,padx=20,pady=5, sticky=tk.E)
    tans.config(height=30,width=20)
    tans.grid(column=9,row=2,columnspan=3, rowspan=30, padx=5,pady=5,sticky=tk.E)
    ba.grid(column=2,row=33,padx=2,pady=5,sticky=tk.E)
    bb.grid(column=3,row=33,padx=2,pady=5,sticky=tk.E)
    bc.grid(column=4,row=33,padx=2,pady=5,sticky=tk.E)
    bd.grid(column=5,row=33,padx=2,pady=5,sticky=tk.E)
    bprev.grid(column=2, row=36, columnspan=2, padx=2, pady=5,sticky=tk.E)
    bnext.grid(column=4, row=36, columnspan=2, padx=2, pady=5,sticky=tk.E)
    Lresult.grid(column=9,row=36,padx=2,pady=5)
    # 创建填空题页面的 Frame 容器，并添加到 Notebook 控件中
    frm3=ttk.Frame(note)
    frm3.pack(fill='both',expand=True)
    note.add(frm3,text="填空题",state="disabled")
    # 在填空题页面中创建 Label 和 ScrolledText 控件用于显示题目，以及 Entry 控件用于输入答案
    Lqus_blank = ttk.Label(frm3, text="题目：")
    tqus_blank = sc.ScrolledText(frm3)
    tqus_blank.config(state="disabled")
    Lanswer = ttk.Label(frm3, text="您的答案：")
    tanswer = ttk.Entry(frm3)

    # 创建提交答案的按钮
    bsubmit_blank = ttk.Button(frm3, text="下一题", command=submit_blank_answer)


    # 创建用于显示用户答题结果的 Label 控件
    Lresult_blank = ttk.Label(frm3)
    Lans_blank = ttk.Label(frm3, text="解析：")
    tans_blank = sc.ScrolledText(frm3)
    tans_blank.config(state="disabled")
    # 使用 grid 布局管理器对填空题页面的控件进行布局
    Lqus_blank.grid(column=0, row=1, padx=20, pady=15, sticky=tk.W)
    Lanswer.grid(column=0, row=33, padx=2, pady=5, sticky=tk.E)
    tanswer.grid(column=1, row=33, padx=2, pady=5, sticky=tk.W)
    bsubmit_blank.grid(column=2, row=33, padx=2, pady=5, sticky=tk.W)
    Lans_blank.grid(column=9, row=1, padx=5, pady=15, sticky=tk.W)
    tqus_blank.config(height=30, width=80)
    tqus_blank.grid(column=0, row=2, columnspan=9, rowspan=30, padx=20, pady=5, sticky=tk.E)
    tans_blank.config(height=30, width=20)
    tans_blank.grid(column=9, row=2, columnspan=3, rowspan=30, padx=5, pady=5, sticky=tk.E)
    Lresult_blank.grid(column=9, row=36, padx=2, pady=5)



    # 创建判断题页面的 Frame 容器，并添加到 Notebook 控件中
    frm4=ttk.Frame(note)
    frm4.pack(fill='both',expand=True)
    note.add(frm4,text="判断题",state="disabled")
    # 在判断题页面中创建 Label 和 ScrolledText 控件用于显示题目，以及 Radiobutton 控件用于选择答案
    Lqus1=ttk.Label(frm4,text="题目：")
    tqus1=sc.ScrolledText(frm4)
    tqus1.config(state="disabled")
    vv=tk.StringVar()# 创建一个字符串变量，用于跟踪 Radiobutton 的选择
    ba1=ttk.Radiobutton(frm4, text="A", variable=vv, value="A",command=radioCa)
    bb1=ttk.Radiobutton(frm4, text="B", variable=vv, value="B",command=radioCa)


    # 创建下一题和上一题的按钮，以及解析 Label 和 ScrolledText 控件
    bnext1=ttk.Button(frm4,text="下一题",command=nextCallpd)
    bprev1 = ttk.Button(frm4, text="上一题", command=prev_pd)
    Lans1=ttk.Label(frm4,text="解析：")
    tans1=sc.ScrolledText(frm4) 
    tans1.config(state="disabled")
    # 为题目文本框添加当获得焦点时自动加载题目的事件绑定
    tqus1.bind("<FocusIn>",load)
    # 创建用于显示用户答题结果的 Label 控件
    Lresult1=tk.Label(frm4)
    # 使用 grid 布局管理器对选择题页面的控件进行布局
    Lqus1.grid(column=0, row=1, padx=20,pady=15, sticky=tk.W)
    Lans1.grid(column=9, row=1, padx=5, pady=15,sticky=tk.W)
    tqus1.config(height=30,width=80)
    tqus1.grid(column=0, row=2, columnspan=9, rowspan=30,padx=20,pady=5, sticky=tk.E)
    tans1.config(height=30,width=20)
    tans1.grid(column=9,row=2,columnspan=3, rowspan=30, padx=5,pady=5,sticky=tk.E)
    ba1.grid(column=2,row=33,padx=2,pady=5,sticky=tk.E)
    bb1.grid(column=3,row=33,padx=2,pady=5,sticky=tk.E)
    bprev1.grid(column=2, row=36, columnspan=2, padx=2, pady=5,sticky=tk.E)
    bnext1.grid(column=4, row=36, columnspan=2, padx=2, pady=5,sticky=tk.E)
    Lresult1.grid(column=9,row=36,padx=2,pady=5)

    # 创建收藏夹页面的 Frame 容器，并添加到 Notebook 控件中
    frm6=ttk.Frame(note)
    frm6.pack(fill='both',expand=True)
    note.add(frm6,text="收藏夹",state="normal")
    
    # 在收藏夹页面中创建 ScrolledText 控件用于显示收藏的题目
    favorite_text = sc.ScrolledText(frm6, height=40, width=100, wrap=tk.WORD)
    favorite_text.pack(fill="both", expand=True, padx=10, pady=10)
    favorite_text.config(state="disabled")

    # 修改选择题页面，添加收藏按钮
    bfavorite = ttk.Button(frm2, text="加入收藏", command=lambda: toggle_favorite("选择题", qus[randselect[indexofselect-1]], ans[randselect[indexofselect-1]], analyze[randselect[indexofselect-1]]))
    bfavorite.grid(column=6, row=36, columnspan=2, padx=2, pady=5, sticky=tk.E)

    # 修改填空题页面，添加收藏按钮
    bfavorite_blank = ttk.Button(frm3, text="加入收藏", command=lambda: toggle_favorite("填空题", blanks[blank_index], blank_ans[blank_index]))
    bfavorite_blank.grid(column=3, row=33, padx=2, pady=5, sticky=tk.W)

    # 修改判断题页面，添加收藏按钮
    bfavorite_pd = ttk.Button(frm4, text="加入收藏", command=lambda: toggle_favorite("判断题", pd[randpd[indexofpd-1]], pd_ans[randpd[indexofpd-1]]))
    bfavorite_pd.grid(column=4, row=33, padx=2, pady=5, sticky=tk.E)

    # 为整个 Notebook 控件添加点击事件，用于处理点击行为
    note.bind("<Button-1>",click)
    mWin.mainloop()
    
if __name__=="__main__":
    main()
