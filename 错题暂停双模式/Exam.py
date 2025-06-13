import openpyxl as px
import random
# 文本文件处理类
class TxtFile:
# 读取小测验游戏说明.txt内容
    @classmethod
    def getGameInfo(cls) :
        s=''
        try:
            f=open("小测验游戏说明.txt")
        except:
            return s
        else:
            while True:
                line=f.readline()
                if not line:
                    break
                s+=line+"\n"
            f.close()
            return s
# 读取Record.txt中内容
    @classmethod
    def getMaxScore(cls) :
        try :
            f=open("Record.txt")
        except: # 文件不存在，第一个人玩家创纪录
            return -1
        else:# 已有旧记录，将得分读出来
            r=f.readlines()
            f.close()
            return int (r[3][3:].strip())

    @classmethod
    def setNewRecord (cls, sno, sname, score):
        f=open("Record.txt",'w')
        f.write("最高得分记录如下: \n")
        f.write('学号：'+sno+'\n')
        f.write('姓名：'+sname+'\n')
        f.write("得分: "+str (score)+"\n")
        f.close ()
# 学生类
class Stu:
    data=0
    sheet=0
    
    def __init__(self):
        Stu.data=px.load_workbook("名单.xlsx")
        Stu.sheet=Stu.data.active

    def getStu(self):
        stu_info=[]
        row_num=Stu.sheet.max_row # 获取最大的行数
        rows=Stu.sheet["A2":"B%d" %row_num]
        for row in rows:# range(1,row_num+1): #行号从1开始计数
            stu_info.append((str(row[0].value),str(row[1].value)))
        return stu_info


import openpyxl as px  
 # 测验试卷类 
class QusAndAns:
    data = 0
    sheet1 = 0
    sheet2 = 0
    sheet3 = 0
    sheet4 = 0
    totalSelect=0
    iselectScore=0
    qus_select=[]
    ans_select=[]
    analyze_select=[]
    qus_blank=[]
    ans_blank=[]
    analyze_blank=[]
    qus_pd=[]
    ans_pd=[]

    

    def __init__(self):
        self.data = px.load_workbook("Quiz.xlsx")  
        self.sheet1 = self.data['选择题']  
        self.sheet2 = self.data['填空题']  
        self.sheet3 = self.data["测验信息"]
        self.sheet4 = self.data['判断题']
        self.cnt_select = self.sheet1.cell(row=1, column=8).value  
        self.cnt_blank = self.sheet2.cell(row=1, column=8).value
        self.cnt_pd = self.sheet4.cell(row=2, column=6).value
        self.examName = self.sheet3.cell(row=1, column=2).value  
        self.examTime = int(self.sheet3.cell(row=2, column=2).value)
# 获取测验名和测验时长
    def getEnameAndEtime(self):
        self.examName=self.sheet3.cell(1,2).value
        self.examTime=int(self.sheet3.cell(2,2).value)
        return(self.examName,self.examTime)
    # 获取选择题总分和每题分数
    def getTotalAndiScore(self):
        QusAndAns.totalSelect=self.sheet1.cell(1,4).value
        QusAndAns.iselectScore=int(self.sheet1.cell(1,6).value)
        return (QusAndAns.totalSelect,QusAndAns.iselectScore)
# 获取选择题
    def getQusOfSelect(self):
        for row in range(3,3+self.cnt_select):
            con=self.sheet1["A%d" %row:'F%d'%row]
            temp1="、"+str (con[0][1].value)+"\n"
            temp2="A. " +str(con[0][2].value)+"\n"+"B. "+str(con[0][3].value)+"\n"+"C. "+str(con[0][4].value)+"\n"+"D. "+str(con[0][5].value)
            QusAndAns.qus_select.append((temp1+temp2))
        return QusAndAns.qus_select
 # 获取选择题答案
    def getAnsOfSelect(self):
        for row in range(3,3+self.cnt_select):
            con=self.sheet1['G%d'%row:'G%d'%row]
            QusAndAns.ans_select.append(str(con[0][0].value))
        return QusAndAns.ans_select
# 获取选择题解析
    def getAnalyzeOfSelect(self):
        for row in range(3,3+self.cnt_select):
            con=self.sheet1["H%d"%row:'H%d'%row]
            QusAndAns.analyze_select.append(str(con[0][0].value))
        return QusAndAns.analyze_select 
 # 获取填空题
    def getQusOfBlank(self):
        # print(self.cnt_blank)
        for row in range(3, 3 + self.cnt_blank):
            con = self.sheet2["B%d" % row:'B%d' % row]
            temp1 = "、" + str(con[0][0].value)
            QusAndAns.qus_blank.append(temp1)
        return QusAndAns.qus_blank
# 获取填空题答案
    def getAnsOfBlank(self):
        for row in range(3, 3 + self.cnt_blank):
            con = self.sheet2['C%d' % row:'C%d' % row]
            QusAndAns.ans_blank.append(str(con[0][0].value))
        return QusAndAns.ans_blank
  # 生成选择题题目的随机顺序
    def getRandQusOfSelect(self):
        randselect=[n for n in range(self.cnt_select)]
        random.shuffle(randselect)
        return randselect

    def getRandAnsOfSelect(self):
        rand=[n for n in range(QusAndAns.ans_select)]
        random.shuffle(rand)
        return rand
   # 生成填空题题目的随机顺序
    def getRandQusOfBlank(self):
        randblank = [n for n in range(self.cnt_blank)]
        random.shuffle(randblank)
        return randblank

# 获取判断题
    def getQusOpd(self):
        for row in range(3,3+self.cnt_pd):
            con1=self.sheet4["A%d" %row:'D%d'%row]
            temp3="、"+str (con1[0][1].value)+"\n"
            temp4="A. " +str(con1[0][2].value)+"\n"+"B. "+str(con1[0][3].value)
            QusAndAns.qus_pd.append((temp3+temp4))
        return QusAndAns.qus_pd

#获取判断题答案
    def getAnsOpd(self):
        for row in range(3,3+self.cnt_pd):
            con=self.sheet4['E%d'%row:'E%d'%row]
            QusAndAns.ans_pd.append(str(con[0][0].value))
        return QusAndAns.ans_pd
#获取判断题解析
    def getRandQusOfpd(self):
        randpd=[n for n in range(self.cnt_pd)]
        random.shuffle(randpd)
        return randpd

